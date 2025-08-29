# services/excel_service.py
"""
Servicio para generación de reportes en formato Excel.
Crea archivos Excel con estadísticas de perfiles y ejecuciones de búsqueda.
"""

# Archivos relacionados: services/profile_service.py

import os
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelService:
    def __init__(self, reports_dir="reports"):
        """
        Inicializa el servicio de Excel

        Args:
            reports_dir (str): Directorio donde guardar los reportes
        """
        self.reports_dir = Path(reports_dir)
        self.ensure_reports_directory()

    def ensure_reports_directory(self):
        """Asegura que el directorio de reportes existe"""
        self.reports_dir.mkdir(exist_ok=True)

    def check_openpyxl_available(self):
        """
        Verifica si openpyxl está disponible

        Returns:
            bool: True si openpyxl está disponible
        """
        return OPENPYXL_AVAILABLE

    def generate_profiles_report(self, profiles_stats, filename=None):
        """
        Genera un reporte de perfiles en Excel

        Args:
            profiles_stats (dict): Estadísticas de perfiles
            filename (str, optional): Nombre del archivo

        Returns:
            str: Ruta del archivo generado

        Raises:
            Exception: Si hay error generando el reporte
        """
        if not OPENPYXL_AVAILABLE:
            raise Exception("openpyxl no esta instalado. Instale con: pip install openpyxl")

        try:
            # Generar nombre de archivo si no se proporciona
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"registro_de_bots_{timestamp}.xlsx"

            filepath = self.reports_dir / filename

            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Registro de Bots"

            # Configurar estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            cell_alignment = Alignment(horizontal="left", vertical="center")
            border_style = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )

            # Título principal
            ws.merge_cells("A1:C1")
            title_cell = ws["A1"]
            title_cell.value = "REGISTRO DE BOTS"
            title_cell.font = Font(bold=True, size=16, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")

            # Fecha del reporte
            ws.merge_cells("A2:C2")
            date_cell = ws["A2"]
            date_cell.value = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            date_cell.font = Font(italic=True)
            date_cell.alignment = Alignment(horizontal="center", vertical="center")

            # Encabezados de columnas
            headers = ["Nombre del Perfil", "Correos Encontrados", "Ultima Ejecucion"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border_style

            # Datos de perfiles
            row = 5
            total_emails_current = 0

            for profile_id, data in profiles_stats.items():
                profile = data.get("profile", {})
                stats = data.get("stats", {})

                profile_name = profile.get("name", "Sin nombre")
                # CORRECCIÓN: Usar current_emails_found en lugar de total_emails_found
                emails_found = stats.get("current_emails_found", stats.get("total_emails_found", 0))
                last_execution = stats.get("last_execution")

                # Formatear fecha de última ejecución
                if last_execution:
                    try:
                        last_exec_dt = datetime.fromisoformat(last_execution.replace('Z', '+00:00'))
                        last_exec_str = last_exec_dt.strftime("%d/%m/%Y %H:%M")
                    except Exception:
                        last_exec_str = "Error de fecha"
                else:
                    last_exec_str = "Nunca"

                # Escribir datos
                cells_data = [profile_name, emails_found, last_exec_str]
                for col, value in enumerate(cells_data, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = value
                    cell.alignment = cell_alignment
                    cell.border = border_style

                total_emails_current += emails_found
                row += 1

            # Fila de totales
            if row > 5:  # Solo si hay datos
                ws.merge_cells(f"A{row}:A{row}")
                total_cell = ws[f"A{row}"]
                total_cell.value = "TOTAL"
                total_cell.font = Font(bold=True)
                total_cell.alignment = Alignment(horizontal="right", vertical="center")
                total_cell.border = border_style

                total_value_cell = ws[f"B{row}"]
                total_value_cell.value = total_emails_current
                total_value_cell.font = Font(bold=True)
                total_value_cell.alignment = cell_alignment
                total_value_cell.border = border_style

                # Celda vacía para la columna de última ejecución
                empty_cell = ws[f"C{row}"]
                empty_cell.border = border_style

            # Ajustar ancho de columnas
            ws.column_dimensions["A"].width = 25
            ws.column_dimensions["B"].width = 18
            ws.column_dimensions["C"].width = 20

            # Añadir fila vacía y resumen
            row += 2
            summary_cell = ws[f"A{row}"]
            summary_cell.value = f"Total de perfiles: {len(profiles_stats)}"
            summary_cell.font = Font(italic=True)

            row += 1
            summary_cell2 = ws[f"A{row}"]
            summary_cell2.value = f"Total de correos encontrados actualmente: {total_emails_current}"
            summary_cell2.font = Font(italic=True)

            # Guardar archivo
            wb.save(filepath)
            wb.close()

            return str(filepath)

        except Exception as e:
            error_msg = self._clean_string(str(e))
            raise Exception(f"Error generando reporte Excel: {error_msg}")

    def generate_detailed_report(self, profiles_stats, filename=None):
        """
        Genera un reporte detallado con múltiples hojas

        Args:
            profiles_stats (dict): Estadísticas de perfiles
            filename (str, optional): Nombre del archivo

        Returns:
            str: Ruta del archivo generado

        Raises:
            Exception: Si hay error generando el reporte
        """
        if not OPENPYXL_AVAILABLE:
            raise Exception("openpyxl no esta instalado. Instale con: pip install openpyxl")

        try:
            # Generar nombre de archivo si no se proporciona
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"reporte_detallado_{timestamp}.xlsx"

            filepath = self.reports_dir / filename

            # Crear workbook
            wb = openpyxl.Workbook()

            # Hoja de resumen
            self._create_summary_sheet(wb, profiles_stats)

            # Hoja de detalles por perfil
            self._create_details_sheet(wb, profiles_stats)

            # Guardar archivo
            wb.save(filepath)
            wb.close()

            return str(filepath)

        except Exception as e:
            error_msg = self._clean_string(str(e))
            raise Exception(f"Error generando reporte detallado: {error_msg}")

    def _create_summary_sheet(self, workbook, profiles_stats):
        """
        Crea la hoja de resumen

        Args:
            workbook: Workbook de openpyxl
            profiles_stats (dict): Estadísticas de perfiles
        """
        ws = workbook.active
        ws.title = "Resumen"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Título
        ws.merge_cells("A1:D1")
        title_cell = ws["A1"]
        title_cell.value = "RESUMEN DE PERFILES"
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Encabezados
        headers = ["Perfil", "Estado", "Ejecuciones", "Correos Encontrados"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Datos
        row = 4
        for profile_id, data in profiles_stats.items():
            profile = data.get("profile", {})
            stats = data.get("stats", {})

            profile_name = profile.get("name", "Sin nombre")
            is_active = "Activo" if profile.get("is_active", True) else "Inactivo"
            executions = stats.get("total_executions", 0)
            # CORRECCIÓN: Usar current_emails_found
            emails_found = stats.get("current_emails_found", stats.get("total_emails_found", 0))

            cells_data = [profile_name, is_active, executions, emails_found]
            for col, value in enumerate(cells_data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center")

            row += 1

        # Ajustar columnas
        for col in ["A", "B", "C", "D"]:
            ws.column_dimensions[col].width = 20

    def _create_details_sheet(self, workbook, profiles_stats):
        """
        Crea la hoja de detalles

        Args:
            workbook: Workbook de openpyxl
            profiles_stats (dict): Estadísticas de perfiles
        """
        ws = workbook.create_sheet(title="Detalles")

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Título
        ws.merge_cells("A1:E1")
        title_cell = ws["A1"]
        title_cell.value = "DETALLES DE PERFILES"
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Encabezados
        headers = ["Perfil", "Criterio Busqueda", "Creado", "Ultima Ejecucion", "Estado"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Datos
        row = 4
        for profile_id, data in profiles_stats.items():
            profile = data.get("profile", {})
            stats = data.get("stats", {})

            profile_name = profile.get("name", "Sin nombre")
            search_title = profile.get("search_title", "Sin criterio")
            created_at = profile.get("created_at", "")
            last_execution = stats.get("last_execution", "Nunca")
            is_active = "Activo" if profile.get("is_active", True) else "Inactivo"

            # Formatear fechas
            if created_at:
                try:
                    created_dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                    created_str = created_dt.strftime("%d/%m/%Y %H:%M")
                except Exception:
                    created_str = created_at
            else:
                created_str = "Sin fecha"

            if last_execution and last_execution != "Nunca":
                try:
                    last_exec_dt = datetime.fromisoformat(last_execution.replace('Z', '+00:00'))
                    last_exec_str = last_exec_dt.strftime("%d/%m/%Y %H:%M")
                except Exception:
                    last_exec_str = last_execution
            else:
                last_exec_str = "Nunca"

            cells_data = [profile_name, search_title, created_str, last_exec_str, is_active]
            for col, value in enumerate(cells_data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center")

            row += 1

        # Ajustar columnas
        column_widths = [25, 30, 18, 18, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

    def create_csv_report(self, profiles_stats, filename=None):
        """
        Genera un reporte en formato CSV como alternativa

        Args:
            profiles_stats (dict): Estadísticas de perfiles
            filename (str, optional): Nombre del archivo

        Returns:
            str: Ruta del archivo generado

        Raises:
            Exception: Si hay error generando el reporte
        """
        try:
            # Generar nombre de archivo si no se proporciona
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"registro_de_bots_{timestamp}.csv"

            filepath = self.reports_dir / filename

            with open(filepath, "w", encoding="utf-8") as f:
                # Encabezados
                f.write("Nombre del Perfil,Correos Encontrados,Ultima Ejecucion\n")

                # Datos
                for profile_id, data in profiles_stats.items():
                    profile = data.get("profile", {})
                    stats = data.get("stats", {})

                    profile_name = self._clean_csv_value(profile.get("name", "Sin nombre"))
                    # CORRECCIÓN: Usar current_emails_found
                    emails_found = stats.get("current_emails_found", stats.get("total_emails_found", 0))
                    last_execution = stats.get("last_execution")

                    # Formatear fecha
                    if last_execution:
                        try:
                            last_exec_dt = datetime.fromisoformat(last_execution.replace('Z', '+00:00'))
                            last_exec_str = last_exec_dt.strftime("%d/%m/%Y %H:%M")
                        except Exception:
                            last_exec_str = "Error de fecha"
                    else:
                        last_exec_str = "Nunca"

                    f.write(f"{profile_name},{emails_found},{last_exec_str}\n")

            return str(filepath)

        except Exception as e:
            error_msg = self._clean_string(str(e))
            raise Exception(f"Error generando reporte CSV: {error_msg}")

    def _clean_csv_value(self, value):
        """
        Limpia un valor para CSV (maneja comas y comillas)

        Args:
            value: Valor a limpiar

        Returns:
            str: Valor limpio para CSV
        """
        if not value:
            return ""

        value_str = str(value)
        # Si contiene coma, envolver en comillas
        if "," in value_str or '"' in value_str:
            # Escapar comillas existentes
            value_str = value_str.replace('"', '""')
            return f'"{value_str}"'

        return value_str

    def get_available_reports(self):
        """
        Obtiene lista de reportes disponibles en el directorio

        Returns:
            list: Lista de archivos de reporte
        """
        try:
            if not self.reports_dir.exists():
                return []

            reports = []
            for file in self.reports_dir.iterdir():
                if file.is_file() and file.suffix.lower() in ['.xlsx', '.csv']:
                    reports.append({
                        "name": file.name,
                        "path": str(file),
                        "size": file.stat().st_size,
                        "modified": file.stat().st_mtime
                    })

            # Ordenar por fecha de modificación (más reciente primero)
            reports.sort(key=lambda x: x["modified"], reverse=True)

            return reports

        except Exception:
            return []

    def _clean_string(self, text):
        """
        Limpia un string de caracteres problemáticos para ASCII

        Args:
            text (str): Texto a limpiar

        Returns:
            str: Texto limpio
        """
        if not text:
            return ""

        try:
            # Reemplazar caracteres problemáticos comunes
            text = text.replace('\xa0', ' ')  # Espacio no-rompible
            text = text.replace('\u2019', "'")  # Apostrofe curvo
            text = text.replace('\u2018', "'")  # Apostrofe curvo
            text = text.replace('\u201c', '"')  # Comilla curva
            text = text.replace('\u201d', '"')  # Comilla curva

            # Codificar y decodificar para limpiar caracteres problemáticos
            return text.encode('ascii', 'ignore').decode('ascii')
        except Exception:
            return str(text)